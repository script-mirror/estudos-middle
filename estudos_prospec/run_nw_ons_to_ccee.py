import datetime as dt
import logging
import os
import shutil
import sys
import time
import pandas as pd
from collections import defaultdict
from datetime import datetime
import zipfile
import openpyxl
from pathlib import Path
from middle.utils import Constants, create_directory, criar_logger
from tqdm import tqdm
from middle.message import send_whatsapp_message
from middle.s3 import (
    handle_webhook_file,
    get_latest_webhook_product,
)
consts = Constants()
sys.path.append(os.path.join(consts.PATH_PROJETOS, "estudos-middle/api_prospec"))
sys.path.append(os.path.join(consts.PATH_PROJETOS, "estudos-middle/update_estudos"))
import run_prospec
from functionsProspecAPI import getStudiesByTag, authenticateProspec, getInfoFromStudy, downloadFileFromDeckV2
from update_newave import NewaveUpdater
# Configure logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

def get_deck_interno():
    logging.info("Authenticating with Prospec API")
    authenticateProspec(consts.API_PROSPEC_USERNAME, consts.API_PROSPEC_PASSWORD)
    logging.info("Fetching studies with tag 'FCF'")
    estudos = getStudiesByTag({'page':1, 'pageSize':3, 'tags':'FCF'})
    logging.debug(f"Found {len(estudos['ProspectiveStudies'])} studies")
    for estudo in estudos['ProspectiveStudies']:
        if estudo['Status'] == 'Concluído':
            prospecId = estudo['Id']
            logging.info(f"Processing study ID: {prospecId}")
            prospecStudy = getInfoFromStudy(prospecId)
            listOfDecks = prospecStudy['Decks']
            logging.debug(f"Found {len(listOfDecks)} decks for study {prospecId}")
            for deck in listOfDecks:
                if deck['Model'] == 'NEWAVE':
                    path = consts.PATH_RESULTS_PROSPEC + '/newave/' + deck['FileName']
                    arrayOfFiles = ['dger.dat', 'vazpast.dat', 'vazoes.dat', 'arquivos.dat', 'confhd.dat']
                    logging.info(f"Downloading deck {deck['FileName']} to {path}")
                    downloadFileFromDeckV2(deck['Id'], consts.PATH_RESULTS_PROSPEC + '/newave/', deck['FileName'], deck['FileName'], arrayOfFiles)
                    time.sleep(15)
                    logging.info(f"Extracting zip folder: {path}")
                    return extract_zip_folder(path, path)

def extract_zip_folder(zip_path: str, folder: str) -> str:
    """Extract a zip folder to the specified path."""
    logging.info(f"Extracting zip file: {folder} to {zip_path}")
    unzip_path: str = os.path.join(zip_path, folder[:-4])
    if not os.path.exists(unzip_path):
        logging.debug(f"Creating extraction directory: {unzip_path}")
        os.makedirs(unzip_path)
    with zipfile.ZipFile(os.path.join(zip_path, folder), 'r') as zip_ref:
        logging.debug(f"Extracting files to: {unzip_path}")
        zip_ref.extractall(unzip_path)
    logging.info(f"Zip extracted to: {unzip_path}")
    return unzip_path

def gtmin_ccee(src):
    """Read GTMIN data from Excel and return a dictionary mapping usid to month and gtmin_agente."""
    logging.info(f"Loading GTMIN data from Excel: {src}")
    try:
        workbook = openpyxl.load_workbook(src)
        sheet = workbook['GTmin_e_CCEE']
        gtmin = defaultdict(dict)
        logging.debug(f"Reading rows from sheet 'GTmin_e_CCEE'")
        for row in tqdm(range(3, len(list(sheet.rows))), desc="Reading Excel rows"):
            try:
                usid = int(sheet.cell(row, 1).value)
                month = sheet.cell(row, 3).value.date()
                gtmin_agente = float(sheet.cell(row, 4).value)
                gtmin[usid][month] = gtmin_agente
                logging.debug(f"Processed row {row}: USID={usid}, Month={month}, GTMIN={gtmin_agente}")
            except (TypeError, ValueError) as e:
                logging.warning(f"Skipping row {row} due to error: {e}")
        logging.info(f"Successfully loaded GTMIN data from {src}")
        return dict(gtmin)
    except Exception as e:
        logging.error(f"Failed to read Excel file {src}: {e}")
        raise

def update_gtmin(arquivo, data, data_deck):
    logger = criar_logger('logging_gtmin.log', os.path.join(os.path.dirname(arquivo), 'logging_gtmin.log'))
    """Update EXPT file with CCEE data and insert TEIFT lines."""
    logger.info(f"Updating EXPT file: {arquivo}")
    with open(arquivo, 'r') as f1:
        expt = f1.readlines()
    rows = ''
    for i in tqdm(range(2, len(expt)), desc="Processing EXPT lines"):
        ln = expt[i]
        try:
            usid = int(ln[:4])
            first_month, first_year = int(ln[20:22]), int(ln[23:27])
            first_date = dt.date(first_year, first_month, 1)
            if 'GTMIN' in ln:
                inflex_ons = float(ln[12:19])
                logger.debug(f"USID: {usid}, Date: {first_date}, Inflex_ONS: {inflex_ons}")
                if first_date.year < 2030:
                    inflex_ccee = data.get(usid, {}).get(first_date, float('inf'))
                    if inflex_ccee < inflex_ons:
                        ln = ln.replace(ln[12:19], f"{inflex_ccee:.2f}".zfill(7).rjust(7))
                        logger.info(f"Updated GTMIN for USID: {str(usid).rjust(3)},  Old value: {str(inflex_ons).rjust(6)},  New value: {str(inflex_ccee).rjust(6)}")
        except (ValueError, IndexError) as e:
            logger.info(f"Invalid line format at line {i+1}: {ln.strip()} - {e}")
            continue
        rows += ln
        if i + 1 < len(expt):
            next_ln = expt[i + 1]
            next_usid = int(next_ln[:4])
            if next_usid != usid:
                next_month = (data_deck.replace(day=28) + dt.timedelta(days=4)).replace(day=1)
                extra_line = f"{str(usid).rjust(4)} TEIFT     0.00 {data_deck.month:>2} {data_deck.year} {next_month.month:>2} {next_month.year}\n"
                rows += extra_line
                logger.debug(f"Added line: {str(usid).rjust(4)} TEIFT     0.00 {data_deck.month:>2} {data_deck.year} {next_month.month:>2} {next_month.year}")
                
    new_expt = expt[0] + expt[1] + rows
    logger.info(f"Writing updated EXPT file: {arquivo}")
    with open(arquivo, 'w') as f:
        for line in new_expt:
            f.write(line)
    logger.info(f"EXPT file updated: {arquivo}")

def update_re(file_path):
    logger = criar_logger('logging_re.log', os.path.join(os.path.dirname(file_path), 'logging_re.log'))
    logger.info(f"Updating RE file: {file_path}")
    with open(file_path, 'r', encoding='latin1') as f:
        lines = f.readlines()

    with open(file_path, 'w', encoding='latin1') as f:
        for line in lines:
            parts = line.strip().split(';')
            if len(parts) >= 2 and parts[1].strip() == '10':
                line = '&' + line
                logger.info(f"Commented line: {line.strip()}")
            f.write(line)
    logger.info(f"RE file updated: {file_path}")

def update_confhd(confhd_path, internal_path):
    logging.info(f"Updating confhd: {confhd_path} with internal: {internal_path}")
    confhd_path = match_file_case_insensitive(confhd_path)
    internal_path = match_file_case_insensitive(internal_path)
    logging.debug(f"Resolved paths: confhd={confhd_path}, internal={internal_path}")

    with open(confhd_path, 'r', encoding='latin1') as f:
        confhd_lines = f.readlines()

    with open(internal_path, 'r', encoding='latin1') as f:
        internal_lines = f.readlines()

    internal_ids = {line.strip().split()[0] for line in internal_lines if line.strip()}
    logging.debug(f"Found {len(internal_ids)} internal IDs")

    new_lines = [
        line for line in confhd_lines
        if line.strip() and line.strip().split()[0] not in internal_ids
    ]
    logging.debug(f"Adding {len(new_lines)} new lines to {internal_path}")

    with open(internal_path, 'a', encoding='latin1') as f:
        for line in new_lines:
            f.write(line)

    logging.info(f"{len(new_lines)} line(s) added to '{internal_path}'.")
    print(f"{len(new_lines)} line(s) added to '{internal_path}'.")

def match_file_case_insensitive(filepath):
    logging.info(f"Matching file case-insensitive: {filepath}")
    directory, target_name = os.path.split(filepath)
    directory = directory or '.'
    for filename in os.listdir(directory):
        if filename.lower() == target_name.lower():
            resolved_path = os.path.join(directory, filename)
            logging.debug(f"Found matching file: {resolved_path}")
            return resolved_path
    logging.error(f"File '{target_name}' not found in '{directory}'.")
    raise FileNotFoundError(f"File '{target_name}' not found in '{directory}'.")

def get_deck_newave(path):
    logging.info(f"Fetching NEWAVE deck to path: {path}")
    preliminar = True
    payload = get_latest_webhook_product(consts.WEBHOOK_DECK_NEWAVE_PRELIMINAR)[0] 
    payload_def = get_latest_webhook_product(consts.WEBHOOK_DECK_NEWAVE_DEFINITIVO)[0] 
    
    if datetime.fromisoformat(payload_def['createdAt'].replace('Z', '+00:00')) > datetime.fromisoformat(payload['createdAt'].replace('Z', '+00:00')):
        preliminar = False
        payload = payload_def 
        logging.debug("Using definitive payload as it is newer")
    else:
        logging.debug("Using preliminary payload")
            
    path_deck_nw = handle_webhook_file(payload, path)
    logging.info(f"Extracting NEWAVE deck zip: {path_deck_nw}")
    path_deck_nw = extract_zip_folder(path_deck_nw, path_deck_nw)
    deck_nw = filter_newave_in_dir(path_deck_nw)
    parts = deck_nw.split('.')[0].split('_')
    data_deck = dt.date(int(parts[2]), int(parts[3]), 1)  
    path_deck_nw = path_deck_nw + '/' + deck_nw
    logging.info(f"Extracting inner NEWAVE deck zip: {path_deck_nw}")
    path_deck_nw = extract_zip_folder(path_deck_nw, path_deck_nw)
    logging.info(f"NEWAVE deck processed: {path_deck_nw}, Date: {data_deck}, Preliminary: {preliminar}")
    return path_deck_nw, data_deck, preliminar

def filter_newave_in_dir(diretorio):
    logging.info(f"Filtering NEWAVE files in directory: {diretorio}")
    arquivos = os.listdir(diretorio)
    arquivos_newave = [arq for arq in arquivos if "newave" in arq.lower()]
    logging.debug(f"Found {len(arquivos_newave)} NEWAVE files: {arquivos_newave}")
    print('DECK ENCONTRADO: ', arquivos_newave[0])
    return arquivos_newave[0]

def get_gtmin(path_in):
    logging.info(f"Fetching GTMIN data to path: {path_in}")
    payload = get_latest_webhook_product(consts.WEBHOOK_NOTAS_TECNICAS)
    df_payload =  pd.DataFrame(payload)
    df_payload =  df_payload.sort_values(by=['periodicidade', 'updatedAt'], ascending=[False, False]).reset_index(drop=True)
    df_payload = df_payload.to_dict('records')
    for arquivo in df_payload:
        path = handle_webhook_file(arquivo, path_in)
        logging.info(f"Extracting GTMIN zip: {path}")
        path = extract_zip_folder(path, path)
        for file in os.listdir(path):
            if file.lower().startswith('gtmin_ccee'):
                logging.info(f"Found GTMIN file: {file} in {path}")
                return path, file
    logging.warning("No GTMIN file found")
    return None, None

def get_expt_sistema_name(path):
    logging.info(f"Searching for EXPT file in: {path}")
    for file in os.listdir(path):
        if file.lower().startswith('expt'):
            expt_path = path + '/' + file
            logging.info(f"Found EXPT file: {expt_path}")
        if file.lower().startswith('sistema'):
            sistema = path + '/' + file
            logging.info(f"Found sistema file: {sistema}")
    return expt_path, sistema


def zip_files(deck_path: str, path_out: str) -> None:
    """Zip only the files in the DC folder, without including subdirectories."""
    logging.info(f"Zipping files from {deck_path} to {path_out}")
    with zipfile.ZipFile(path_out, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file in os.listdir(deck_path):
            file_path: str = os.path.join(deck_path, file)
            if os.path.isfile(file_path):
                logging.debug(f"Adding file to zip: {file_path}")
                zipf.write(file_path, arcname=file)
    logging.info(f"Zip file created: {path_out}")

def delete_file(diretorio, lista_nomes):
    logging.info(f"Deleting files from {diretorio}: {lista_nomes}")
    arquivos_no_dir = os.listdir(diretorio)
    arquivos_mapeados = {arq.lower(): arq for arq in arquivos_no_dir}

    for nome in lista_nomes:
        nome_lower = nome.lower()
        if nome_lower in arquivos_mapeados:
            caminho_completo = os.path.join(diretorio, arquivos_mapeados[nome_lower])
            if os.path.isfile(caminho_completo):
                try:
                    os.remove(caminho_completo)
                    logging.info(f"Deleted: {caminho_completo}")
                    print(f"Deletado: {caminho_completo}")
                except Exception as e:
                    logging.error(f"Error deleting {caminho_completo}: {e}")
                    print(f"Erro ao deletar {caminho_completo}: {e}")
            else:
                logging.warning(f"Not a file: {caminho_completo}")
                print(f"Não é um arquivo: {caminho_completo}")
        else:
            logging.warning(f"File not found (case-insensitive): {nome}")
            print(f"🔍 Não encontrado (ignorando case): {nome}")

def execute_prospec(deck_path: str, deck_name: str):
    logging.info(f"Executing Prospec with deck: {deck_name}, path: {deck_path}")
    params = {}
    params['deck'] = f"{deck_name}.zip"
    params['path_deck'] = deck_path + '/'
    params['tag'] = 'FCF'
    params['aguardar_fim'] = False
    params['apenas_email'] = False
    params['back_teste'] = True
    logging.debug(f"Prospec parameters: {params}")
    run_prospec.main(params)
    logging.info("Prospec execution completed")


def update_wind(path):
    updater = NewaveUpdater()
    params = {}
    params['produto'] = 'EOLICA'
    params['id_estudo'] = None
    params['path_download'] = create_directory(consts.PATH_RESULTS_PROSPEC, 'update_decks/' + params['produto']) + '/'
    params['path_out'] = create_directory(consts.PATH_RESULTS_PROSPEC, 'update_decks/' + params['produto']) + '/'
    updater.update_eolica(params,path)

def main():
    logging.info("Starting main execution")
    PATH_BASE = create_directory(consts.PATH_ARQUIVOS, 'decks/newave/ons')
    logging.info(f"Base path created: {PATH_BASE}")
            
    path_deck_nw, data_deck, preliminar = get_deck_newave(PATH_BASE)
    logging.info(f"NEWAVE deck retrieved: {path_deck_nw}, Date: {data_deck}, Preliminary: {preliminar}")
    
    path_gtmin, gtmin_file = get_gtmin(PATH_BASE)
    logging.info(f"GTMIN data retrieved: {path_gtmin}/{gtmin_file}")
    
    path_expt, path_sistema = get_expt_sistema_name(path_deck_nw)
    logging.info(f"EXPT file: {path_expt}")
    
    deck_name = 'NW' + data_deck.strftime('%Y%m') + '_ONS-TO-CCEE.zip'
    att_eolica  = 'Não Atualizada'
    att_vazpast = 'Não Atualizada'
    att_confhd  = 'Não Atualizada'
    if preliminar:
        deck_name = 'NW' + data_deck.strftime('%Y%m') + '_ONS-TO-CCEE_PRELIMINAR.zip'
        logging.info(f"Using preliminary deck name: {deck_name}")
        PATH_NW_INTERNO = get_deck_interno()
        logging.info(f"Internal deck retrieved: {PATH_NW_INTERNO}")
        update_confhd(path_deck_nw + "/confhd.dat", PATH_NW_INTERNO + "/confhd.dat")
        delete_file(path_deck_nw, ['DGER.DAT', 'VAZOES.DAT', 'VAZPAST.DAT', 'ARQUIVOS.DAT', 'CONFHD.DAT'])
        logging.info("Copying internal files to deck path")
        shutil.copy(PATH_NW_INTERNO + '/dger.dat',     path_deck_nw + '/dger.dat')
        shutil.copy(PATH_NW_INTERNO + '/vazoes.dat',   path_deck_nw + '/vazoes.dat')
        shutil.copy(PATH_NW_INTERNO + '/vazpast.dat',  path_deck_nw + '/vazpast.dat')
        shutil.copy(PATH_NW_INTERNO + '/arquivos.dat', path_deck_nw + '/arquivos.dat')
        shutil.copy(PATH_NW_INTERNO + '/confhd.dat',   path_deck_nw + '/confhd.dat')
        update_wind([path_sistema])
        att_eolica  = 'Atualizada'
        att_vazpast = 'Atualizada'
        att_confhd  = 'Atualizada'
    else:
        logging.info(f"Using definitive deck name: {deck_name}")
    
    update_re(path_deck_nw + "/restricao-eletrica.csv")
    update_gtmin(path_expt, gtmin_ccee(path_gtmin + '/' + gtmin_file), data_deck)
    zip_files(path_deck_nw, PATH_BASE + '/' + deck_name)
    logging.info(f"Sending WhatsApp message with deck: {deck_name}")
    send_whatsapp_message(consts.WHATSAPP_GILSEU, 'Deck Newave ONS to CCEE\nGTmin: ' + gtmin_file +\
        f"\nEolica: {att_eolica}" +f"\nvazpast.dat: {att_vazpast}"+f"\nconfhd.dat: {att_confhd}" ,
                          PATH_BASE + '/' + deck_name)
    execute_prospec(PATH_BASE, deck_name.split('.')[0])
    logging.info("Main execution completed")

if __name__ == "__main__":
    main()