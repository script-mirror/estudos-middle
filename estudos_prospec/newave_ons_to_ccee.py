import datetime as dt
import logging
import os
import shutil
import sys
from collections import defaultdict
from datetime import datetime, date, timedelta
import zipfile
from zipfile import ZipFile, ZIP_DEFLATED
from typing import List, Dict, Any, Optional, Tuple
import openpyxl
from pathlib import Path
from middle.utils.constants import Constants 
from tqdm import tqdm
from middle.utils import SemanaOperativa
consts = Constants()
sys.path.append(os.path.join(consts.PATH_PROJETOS, "estudos-middle/api_prospec"))
import run_prospec
from functionsProspecAPI import  getStudiesByTag, authenticateProspec, getInfoFromStudy, downloadFileFromDeckV2

# Configure logging
logging.basicConfig(level=logging.ERROR, format='%(asctime)s - %(levelname)s - %(message)s')

def create_directory(base_path: str, sub_path: str) -> Path:
        full_path = Path(base_path) / sub_path
        try: os.remove(full_path) if os.path.exists(full_path) else None
        except: pass
        try: shutil.rmtree(full_path) if os.path.exists(full_path) else None
        except: pass
        full_path.mkdir(parents=True, exist_ok=True)
        return full_path.as_posix()  

# Define directory paths
os.makedirs(consts.PATH_ARQUIVOS,exist_ok=True)
PATH_BASE = create_directory(consts.PATH_ARQUIVOS, 'decks/nenwave/ons')
PATH_BASE
GTMIN = PATH_BASE + "GTMIN_CCEE_082025_preliminar.xlsx"


def get_deck_interno():
    authenticateProspec(consts.API_PROSPEC_USERNAME, consts.API_PROSPEC_PASSWORD)
    estudos = getStudiesByTag({'page':1, 'pageSize':3, 'tags':'FCF'})
    for estudo in estudos['ProspectiveStudies']:
        if estudo['Status'] == 'Concluído':
            prospecId = estudo['Id']
            prospecStudy = getInfoFromStudy(prospecId)
            listOfDecks  = prospecStudy['Decks']
            for deck in listOfDecks:
                if deck['Model'] == 'NEWAVE':
                    path = consts.PATH_RESULTS_PROSPEC + '/newave/' + deck['FileName']
                    arrayOfFiles = ['dger.dat', 'vazpast.dat', 'vazoes.dat','arquivos.dat', 'confhd.dat']
                    downloadFileFromDeckV2(deck['Id'],consts.PATH_RESULTS_PROSPEC + '/newave/', deck['FileName'], deck['FileName'],arrayOfFiles)
                    return extract_zip_folder(path, path)

def extract_zip_folder(zip_path: str, folder: str) -> str:
    """Extract a zip folder to the specified path."""
    unzip_path: str = os.path.join(zip_path, folder[:-4])
    if not os.path.exists(unzip_path):
        os.makedirs(unzip_path)
    with zipfile.ZipFile(os.path.join(zip_path, folder), 'r') as zip_ref:
        zip_ref.extractall(unzip_path)
    return unzip_path


def gtmin_ccee(src):
    """Read GTMIN data from Excel and return a dictionary mapping usid to month and gtmin_agente."""
    logging.info(f"Loading data from {src}")
    try:
        workbook = openpyxl.load_workbook(src)
        sheet = workbook['GTmin_e_CCEE']
        gtmin = defaultdict(dict)
        for row in tqdm(range(3, len(list(sheet.rows))), desc="Reading Excel rows"):
            try:
                usid = int(sheet.cell(row, 1).value)
                month = sheet.cell(row, 3).value.date()
                gtmin_agente = float(sheet.cell(row, 4).value)
                gtmin[usid][month] = gtmin_agente
            except (TypeError, ValueError) as e:
                logging.warning(f"Skipping row {row} due to error: {e}")
        return dict(gtmin)
    except Exception as e:
        logging.error(f"Failed to read Excel file {src}: {e}")
        raise

def update(arquivo, data, data_deck):
    """Update EXPT file with CCEE data and insert TEIFT lines."""
    expt = arquivo.readlines()
    rows = ''
    for i in tqdm(range(2, len(expt)), desc="Processing EXPT lines"):
        ln = expt[i]
        try:
            usid = int(ln[:4])
            first_month, first_year = int(ln[20:22]), int(ln[23:27])
            first_date = dt.date(first_year, first_month, 1)
            if 'GTMIN' in ln:
                inflex_ons = float(ln[12:19])
                logging.debug(f"USID: {usid}, Date: {first_date}, Inflex_ONS: {inflex_ons}")
                if first_date.year < 2030:
                    inflex_ccee = data.get(usid, {}).get(first_date, float('inf'))
                    if inflex_ccee < inflex_ons:
                        ln = ln.replace(ln[12:19], f"{inflex_ccee:.2f}".zfill(7).rjust(7))
        except (ValueError, IndexError) as e:
            logging.error(f"Invalid line format at line {i+1}: {ln.strip()} - {e}")
            continue
        rows += ln
        if i + 1 < len(expt):
            next_ln = expt[i + 1]
            next_usid = int(next_ln[:4])
            if next_usid != usid:
                next_month = (data_deck.replace(day=28) + dt.timedelta(days=4)).replace(day=1)
                extra_line = f"{str(usid).rjust(4)} TEIFT     0.00 {data_deck.month:>2} {data_deck.year} {next_month.month:>2} {next_month.year}\n"
                rows += extra_line
    return expt[0] + expt[1] + rows

def extract_inner_zip(outer_zip):
    """Extract inner ZIP and return its handle and data_deck date."""
    infolist = outer_zip.infolist()
    if not infolist or not infolist[0].filename.endswith('.zip'):
        raise ValueError("Produtos.zip does not contain a valid inner ZIP file")
    inner_zip = ZipFile(outer_zip.open(infolist[0], mode="r"))
    try:
        parts = inner_zip.filename.split('_')
        data_deck = dt.date(int(parts[2]), int(parts[3]), 1)
    except (IndexError, ValueError):
        raise ValueError(f"Cannot parse date from filename: {inner_zip.filename}")
    return inner_zip, data_deck

def extract_and_find_expt(inner_zip, data_deck, output_dir):
    """Extract inner ZIP to output directory and return EXPT file path."""
    extract_to = os.path.join(output_dir, inner_zip.filename.replace(".zip", ""))
    os.makedirs(extract_to, exist_ok=True)
    inner_zip.extractall(extract_to)
    logging.info(f"Extracted files to: {extract_to}")
    expt_files = [f for f in os.listdir(extract_to) if "EXPT" in f]
    if not expt_files:
        raise FileNotFoundError("No EXPT file found in the extracted directory")
    return os.path.join(extract_to, expt_files[0]), extract_to

def process_expt(expt_path, ccee_data, data_deck):
    """Process EXPT file and return updated content."""
    logging.info(f"Reading EXPT file: {expt_path}")
    with open(expt_path, 'r') as expt_ons:
        logging.info("Updating EXPT data")
        return update(expt_ons, ccee_data, data_deck)

def create_ccee_output(inner_zip, expt_ccee, extract_to, data_deck, output_dir):
    """Create CCEE output directory and ZIP file."""
    extract_to_ccee = extract_to.replace('deck_newave', 'NW_ONS-TO-CCEE')
    os.makedirs(extract_to_ccee, exist_ok=True)
    inner_zip.extractall(extract_to_ccee)
    logging.info(f"Extracted files to CCEE directory: {extract_to_ccee}")
    
    expt_ccee_path = os.path.join(extract_to_ccee, "EXPT.DAT")
    logging.info(f"Writing updated EXPT file: {expt_ccee_path}")
    with open(expt_ccee_path, "w") as f:
        f.write(expt_ccee)
    
    output_zip = os.path.join(output_dir, "NW_CCEE.zip")
    logging.info(f"Creating output ZIP: {output_zip}")
    with ZipFile(output_zip, "w", compression=ZIP_DEFLATED) as deck_newave_ccee:
        for root, _, files in os.walk(extract_to_ccee):
            for fl in files:
                full_path = os.path.join(root, fl)
                arcname = os.path.relpath(full_path, os.path.dirname(extract_to_ccee))
                deck_newave_ccee.write(full_path, arcname)
                
def update_re(file_path):

    with open(file_path, 'r',  encoding='latin1') as f:
        lines = f.readlines()

    with open(file_path, 'w',  encoding='latin1') as f:
        for line in lines:
            parts = line.strip().split(';')
            if len(parts) >= 2 and parts[1].strip() == '10':
                line = '&'+line
            f.write(line)

def update_confhd(confhd_path, internal_path):

    with open(confhd_path, 'r', encoding='latin1') as f1:
        confhd_lines = f1.readlines()

    with open(internal_path, 'r', encoding='latin1') as f2:
        internal_lines = f2.readlines()

    internal_ids = {line.strip().split()[0] for line in internal_lines if line.strip()}

    new_lines = [
        line for line in confhd_lines
        if line.strip() and line.strip().split()[0] not in internal_ids
    ]

    with open(internal_path, 'a', encoding='latin1') as f:
        for line in new_lines:
            f.write(line)

    print(f"{len(new_lines)} line(s) added to '{internal_path}'.")
    
    
def update_confhd(confhd_path, internal_path):
 

    confhd_path = match_file_case_insensitive(confhd_path)
    internal_path = match_file_case_insensitive(internal_path)

    with open(confhd_path, 'r', encoding='latin1') as f:
        confhd_lines = f.readlines()

    with open(internal_path, 'r', encoding='latin1') as f:
        internal_lines = f.readlines()

    internal_ids = {line.strip().split()[0] for line in internal_lines if line.strip()}

    new_lines = [
        line for line in confhd_lines
        if line.strip() and line.strip().split()[0] not in internal_ids
    ]

    with open(internal_path, 'a', encoding='latin1') as f:
        for line in new_lines:
            f.write(line)

    print(f"{len(new_lines)} line(s) added to '{internal_path}'.")


def match_file_case_insensitive(filepath):
  
    directory, target_name = os.path.split(filepath)
    directory = directory or '.'  # handle empty path (current dir)

    for filename in os.listdir(directory):
        if filename.lower() == target_name.lower():
            return os.path.join(directory, filename)
    raise FileNotFoundError(f"File '{target_name}' not found in '{directory}'.")

# Example usage:
            
def main():
    """Main function to process NEWAVE deck and create CCEE output."""
    update_re("C:/projetos/arquivos/restricao-eletrica.csv")
    update_confhd("C:/projetos/arquivos/confhd.dat", "C:/projetos/arquivos/CONFHD_interno.DAT")
    
    
    try:
        
        PATH_NW_INTERNO = get_deck_interno()
        data_rv = SemanaOperativa(datetime.today())
        dt_newave = data_rv.first_day_of_month + timedelta(days=6)      
        nome_deck =  '/NW' + dt_newave.strftime('%Y%m')
        shutil.copy(PATH_NW_INTERNO + '/dger.dat',    PATH_BASE + '/dger.dat')
        shutil.copy(PATH_NW_INTERNO + '/vazoes.dat',  PATH_BASE + '/vazoes.dat')
        shutil.copy(PATH_NW_INTERNO + '/vazpast.dat', PATH_BASE + '/vazpast.dat')
        update_re(PATH_BASE + "/restricao-eletrica.csv")
        update_confhd(PATH_BASE + "/confhd.dat", PATH_NW_INTERNO + "/confhd.dat")
        shutil.copy(PATH_NW_INTERNO + '/confhd.dat', PATH_BASE + '/confhd.dat')
        
        # Define directory paths
        os.makedirs(consts.PATH_ARQUIVOS,exist_ok=True)
        BASE_DIR = "C:/Dev/12 - ConverteNW"
        INPUT_DIR = os.path.join(BASE_DIR, "input")
        OUTPUT_DIR = os.path.join(BASE_DIR, "output")
        ZIP_PATH = os.path.join(INPUT_DIR, "Produtos.zip")
        GTMIN = os.path.join(INPUT_DIR, "GTMIN_CCEE_082025_preliminar.xlsx")

        
        with ZipFile(ZIP_PATH, mode="r") as outer_zip:
            inner_zip, data_deck = extract_inner_zip(outer_zip)
            with inner_zip:
                expt_path, extract_to = extract_and_find_expt(inner_zip, data_deck, OUTPUT_DIR)
                ccee_data = gtmin_ccee(GTMIN)
                expt_ccee = process_expt(expt_path, ccee_data, data_deck)
                create_ccee_output(inner_zip, expt_ccee, extract_to, data_deck, OUTPUT_DIR)
        
        logging.info("Processing completed successfully")
    except Exception as e:
        logging.error(f"Error during execution: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
    sys.exit(0)